from __future__ import annotations
import sqlite3
import csv
from pathlib import Path
from jwbuddy.tools.base import BaseTool, ToolSpec, ToolResult


class QueryFileTool(BaseTool):
    """文件查询工具：对上传的 Excel/CSV 文件执行 SQL 查询"""

    def __init__(self, upload_dir: str = "data/uploads"):
        self.upload_dir = Path(upload_dir)
        self._schemas: dict[str, list[str]] = {}  # Cache column info per file

    @property
    def spec(self) -> ToolSpec:
        return ToolSpec(
            name="query_file",
            description=(
                "对上传的 Excel/CSV 文件执行 SQL 分析查询。"
                "重要：必须先了解文件有哪些列名和样例数据，再编写带 GROUP BY/ORDER BY/聚合函数的分析型 SQL。"
                "示例流程："
                "1) 先用 sql='SELECT * FROM data LIMIT 3' 查看列名和样例数据；"
                "2) 根据列名编写分析型SQL，如："
                "SELECT 市州, SUM(活跃用户数) as 总活跃 FROM data WHERE 月份='5月' GROUP BY 市州 ORDER BY 总活跃 DESC。"
                "SQL 中列名必须与文件中的列名完全一致。表名固定为 data。"
            ),
            parameters={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "上传文件的文件名",
                    },
                    "sql": {
                        "type": "string",
                        "description": (
                            "要执行的 SQL 查询语句（SELECT 开头），表名为 data。"
                            "先了解列名，再写分析型 SQL（GROUP BY/ORDER BY/聚合函数）。"
                            "列名与文件第一行一致，在 SQL 中用双引号包裹含空格的列名。"
                        ),
                    },
                },
                "required": ["file_path", "sql"],
            },
        )

    async def execute(self, **kwargs) -> ToolResult:
        file_path = kwargs.get("file_path", "")
        sql = kwargs.get("sql", "").strip()

        forbidden = ["DROP", "DELETE", "INSERT", "UPDATE", "ALTER", "CREATE", "PRAGMA", "ATTACH", "DETACH"]
        upper_sql_stripped = sql.upper().lstrip()
        for kw in forbidden:
            if kw in upper_sql_stripped:
                return ToolResult(success=False, error=f"禁止的操作: {kw}")

        if not upper_sql_stripped.startswith("SELECT"):
            return ToolResult(success=False, error="仅允许 SELECT 查询")

        path = Path(file_path)
        if ".." in path.parts:
            return ToolResult(success=False, error="文件路径不允许")
        if not path.is_absolute():
            path = (self.upload_dir / path).resolve()
        if not path.exists():
            return ToolResult(success=False, error=f"文件不存在: {file_path}")

        try:
            rows, columns = self._load_file(path)
        except Exception as e:
            return ToolResult(success=False, error=f"读取文件失败: {e}")
        if not rows:
            return ToolResult(success=False, error="文件中没有数据")

        # Cache column names for reference
        self._schemas[file_path] = columns

        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        try:
            clean_cols = [c.replace(" ", "_").replace("-", "_").replace(".", "_") for c in columns]
            col_defs = ", ".join(f'"{c}" TEXT' for c in clean_cols)
            conn.execute(f'CREATE TABLE "data" ({col_defs})')
            placeholders = ", ".join("?" * len(clean_cols))
            conn.executemany(f'INSERT INTO "data" VALUES ({placeholders})', rows)

            mapped_sql = sql
            for orig, clean in zip(columns, clean_cols):
                if orig != clean:
                    mapped_sql = mapped_sql.replace(orig, clean)

            cur = conn.execute(mapped_sql)
            result_rows = [dict(row) for row in cur.fetchall()]
            conn.close()

            summary = f"文件: {path.name} | 列: {', '.join(columns)} | 总行数: {len(rows)} | 查询结果: {len(result_rows)} 行"

            return ToolResult(
                success=True,
                data={
                    "rows": result_rows[:100],
                    "columns": list(result_rows[0].keys()) if result_rows else columns,
                    "total_rows": len(result_rows),
                    "summary": summary,
                },
                format="table",
                text=summary,
            )
        except Exception as e:
            conn.close()
            return ToolResult(success=False, error=f"查询失败: {e}。可用列名: {', '.join(columns)}")

    def _load_file(self, path: Path) -> tuple[list[tuple], list[str]]:
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not all_rows:
                return [], []
            columns = [str(c or "") for c in all_rows[0]]
            data = [tuple(str(v or "") if v is not None else "" for v in row) for row in all_rows[1:]]
            return data, columns
        elif suffix == ".csv":
            with open(path, newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                all_rows = list(reader)
            if not all_rows:
                return [], []
            columns = all_rows[0]
            data = [tuple(row) for row in all_rows[1:]]
            return data, columns
        else:
            raise ValueError(f"不支持的文件格式: {suffix}")
