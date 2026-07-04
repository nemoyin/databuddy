"""Tests for QueryFileTool (TC-7.1 ~ TC-7.9)."""
from __future__ import annotations

import csv
import io
import os
import tempfile
import pytest
from pathlib import Path
from jwbuddy.tools.query_file import QueryFileTool


def _make_xlsx(rows: list[tuple], filename: str = "test.xlsx") -> str:
    """Helper: create a real .xlsx file from row data, return path."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    tmpdir = tempfile.mkdtemp()
    filepath = os.path.join(tmpdir, filename)
    wb.save(filepath)
    return filepath


def _make_csv(rows: list[list[str]], filename: str = "test.csv", encoding: str = "utf-8") -> str:
    """Helper: create a CSV file from row data, return path."""
    tmpdir = tempfile.mkdtemp()
    filepath = os.path.join(tmpdir, filename)
    with open(filepath, "w", newline="", encoding=encoding) as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow(row)
    return filepath


@pytest.mark.asyncio
async def test_query_explore_structure():
    """TC-7.1: Explore data structure with SELECT * FROM data LIMIT 3."""
    filepath = _make_xlsx([
        ("Name", "Age", "City"),
        ("Alice", 30, "Beijing"),
        ("Bob", 25, "Shanghai"),
        ("Charlie", 35, "Shenzhen"),
        ("Diana", 28, "Beijing"),
    ])
    tool = QueryFileTool(upload_dir=os.path.dirname(filepath))

    result = await tool.execute(
        file_path=filepath,
        sql="SELECT * FROM data LIMIT 3",
    )

    assert result.success
    assert result.data["total_rows"] == 3
    assert len(result.data["rows"]) == 3
    assert "Name" in result.data["columns"]


@pytest.mark.asyncio
async def test_query_aggregation():
    """TC-7.2: Aggregation with GROUP BY and ORDER BY."""
    filepath = _make_xlsx([
        ("City", "Sales"),
        ("Beijing", 100),
        ("Shanghai", 200),
        ("Beijing", 150),
        ("Shenzhen", 300),
    ])
    tool = QueryFileTool(upload_dir=os.path.dirname(filepath))

    result = await tool.execute(
        file_path=filepath,
        sql="SELECT City, SUM(Sales) as TotalSales FROM data GROUP BY City ORDER BY TotalSales DESC",
    )

    assert result.success
    rows = result.data["rows"]
    assert len(rows) == 3
    # First row should be Shenzhen with 300
    assert rows[0]["City"] == "Shenzhen"
    assert rows[0]["TotalSales"] in (300, "300")


@pytest.mark.asyncio
async def test_query_where_filter():
    """TC-7.3: WHERE filtering returns filtered data."""
    filepath = _make_xlsx([
        ("Product", "Price", "InStock"),
        ("A", 10, "Yes"),
        ("B", 20, "No"),
        ("C", 15, "Yes"),
    ])
    tool = QueryFileTool(upload_dir=os.path.dirname(filepath))

    result = await tool.execute(
        file_path=filepath,
        sql="SELECT * FROM data WHERE InStock='Yes'",
    )

    assert result.success
    assert result.data["total_rows"] == 2
    for row in result.data["rows"]:
        assert row["InStock"] == "Yes"


@pytest.mark.asyncio
async def test_query_forbid_drop():
    """TC-7.4: Forbidden DROP operation returns error."""
    filepath = _make_xlsx([("A",), (1,)])
    tool = QueryFileTool(upload_dir=os.path.dirname(filepath))

    result = await tool.execute(
        file_path=filepath,
        sql="DROP TABLE data",
    )

    assert not result.success
    assert "禁止" in result.error
    assert "DROP" in result.error


@pytest.mark.asyncio
async def test_query_forbid_insert():
    """TC-7.5: Forbidden INSERT operation returns error."""
    filepath = _make_xlsx([("A",), (1,)])
    tool = QueryFileTool(upload_dir=os.path.dirname(filepath))

    result = await tool.execute(
        file_path=filepath,
        sql="INSERT INTO data VALUES (1, 'test')",
    )

    assert not result.success
    # INSERT is caught by forbidden keyword check (more specific than SELECT-only)
    assert "禁止" in result.error
    assert "INSERT" in result.error


@pytest.mark.asyncio
async def test_query_nonexistent_file():
    """TC-7.6: Non-existent file returns error."""
    tool = QueryFileTool(upload_dir="/tmp/jwb_nonexistent")

    result = await tool.execute(
        file_path="nonexistent_file.xlsx",
        sql="SELECT * FROM data LIMIT 1",
    )

    assert not result.success
    assert "文件不存在" in result.error


@pytest.mark.asyncio
async def test_query_path_traversal():
    """TC-7.7: Path traversal attempt is blocked."""
    tool = QueryFileTool(upload_dir="/tmp/jwb_test")

    result = await tool.execute(
        file_path="../etc/passwd",
        sql="SELECT * FROM data LIMIT 1",
    )

    assert not result.success
    assert "文件路径不允许" in result.error


@pytest.mark.asyncio
async def test_query_multi_sheet_excel():
    """TC-7.8: Multi-sheet Excel reads active sheet correctly."""
    from openpyxl import Workbook
    wb = Workbook()
    # Sheet 1 (default active)
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["ID", "Value"])
    ws1.append([1, "First"])
    # Sheet 2
    ws2 = wb.create_sheet("Details")
    ws2.append(["Name", "Score"])
    ws2.append(["Alice", 95])

    tmpdir = tempfile.mkdtemp()
    filepath = os.path.join(tmpdir, "multi_sheet.xlsx")
    wb.save(filepath)

    tool = QueryFileTool(upload_dir=tmpdir)
    result = await tool.execute(
        file_path=filepath,
        sql="SELECT * FROM data LIMIT 1",
    )

    assert result.success
    # Should read the active sheet (Sheet 1 by default = "Summary")
    assert "ID" in result.data["columns"]


@pytest.mark.asyncio
async def test_query_csv_utf8_bom():
    """TC-7.9: CSV with UTF-8 BOM is parsed correctly."""
    filepath = _make_csv(
        [["Name", "Age"], ["Alice", "30"]],
        filename="bom_test.csv",
        encoding="utf-8-sig",
    )
    tool = QueryFileTool(upload_dir=os.path.dirname(filepath))

    result = await tool.execute(
        file_path=filepath,
        sql="SELECT * FROM data",
    )

    assert result.success
    # Column names should NOT contain BOM character
    assert result.data["columns"][0] == "Name"
    assert result.data["rows"][0]["Name"] == "Alice"
